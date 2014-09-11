#!/usr/bin/env python

import sys
import os
import argparse
import blatfilter
import traceback

from pysam import Tabixfile
from re import search
from collections import defaultdict as dd
from collections import Counter
from openpyxl import Workbook


columns = ['Chr',
           'Left_Position',
           'Right_Position',
           'Pos_Conf',
           'Strand',
           'Strand_Conf',
           'Class',
           'Family',
           'Family_Conf',
           'Ref_TE',
           'Elt_Length',
           'TE_Min_Position',
           'TE_Max_Position',
           'TSD',
           'TSD_Length',
           'DEL',
           'DEL_Length',
           'Left_Snip',
           'Right_Snip',
           'Left_Consensus',
           'Right_Consensus',
           'Left_Support',
           'Right_Support',
           'Total_Support',
           'Libraries',
           'Library_List',
           'Mapscore',
           'Known_Nonref',
           'Genes',
           'Regulation']


def vcf2tld(args, invcf, consfasta):
    ''' convert selected VCF annotations to two-level dictionary '''
    current_row = 0

    # two-level dictionary
    tld = dd(dict)

    # read consensus
    cons = []
    seq  = []

    with open(consfasta, 'r') as consfa:
        for line in consfa:
            if line.startswith('>'):
                cons.append(line.strip().lstrip('>'))
            else:
                seq.append(line.strip())

    consdict = dict(zip(cons, seq))

    sys.stderr.write("build consensus dict of " + str(len(consdict)) + " sequences from " + consfasta + "\n")

    with open(invcf, 'r') as vcf:
        for line in vcf:
            if not line.startswith('#'):
                chrom, pos, id, ref, alt, qual, filter, info, format, sample = line.strip().split('\t') 

                infodict = {}
                for i in info.split(';'):
                    if search('=', i):
                        key, val = i.split('=')
                        infodict[key] = val
                    else:
                        infodict[i] = True

                formatdict = dict([(key, val) for key, val in zip(format.split(':'), sample.split(':'))])

                if filter == 'PASS':

                    start = int(pos)
                    end   = 'NA'
                    if 'END' in infodict:
                        end = int(infodict['END'])
                        if start > end:
                            start, end = end, start

                    alt = alt.split(':')[-1].strip('>')
                    strand = 'NA'
                    strandconf = 0.0

                    if int(formatdict['FWD']) > int(formatdict['REV']):
                        strand = '+'
                        strandconf = float(formatdict['FWD']) / (float(formatdict['FWD']) + float(formatdict['REV']))

                    if int(formatdict['FWD']) < int(formatdict['REV']):
                        strand = '-'
                        strandconf = float(formatdict['REV']) / (float(formatdict['FWD']) + float(formatdict['REV']))

                    # build family info
                    famvec  = []
                    for rawfam in formatdict['TEALIGN'].split(','):
                        fam, count = rawfam.split('|')
                        for _ in range(int(count)):
                            famvec.append(fam)

                    famtotal = len([f for f in famvec if f != 'POLYA'])
                    family, famcount = Counter(famvec).most_common(1)[0]

                    if family == 'POLYA':
                        family, famcount = Counter(famvec).most_common(2)[1]

                    famconf = float(famcount)/float(famtotal)

                    eltlen  = 'NA'
                    if 'TELEN' in infodict:
                        eltlen = infodict['TELEN']


                    teminpos = formatdict['TEMINPOS']
                    temaxpos = formatdict['TEMAXPOS'] 

                    readcount = formatdict['RC']

                    lsupport = 0
                    rsupport = 0
                    for b in formatdict['BREAKS'].split(','):
                        bloc, count = map(int, b.split('|'))
                        if start == bloc:
                            lsupport += count
                        if start != end and end == bloc:
                            rsupport += count 

                    support = lsupport + rsupport
                    posconf = float(support)/float(readcount)

                    TSD = 'N'
                    DEL = 'N'
                    tlen = 0
                    dlen = 0

                    if 'MECH' in infodict:
                        if infodict['MECH'] == 'TSD':
                            TSD = 'Y'
                            tlen = infodict['TSDLEN']

                        if infodict['MECH'] == 'Deletion':
                            DEL = 'Y'
                            dlen = infodict['TSDLEN']

                    rightsnip = formatdict['POSBREAKSEQ']
                    leftsnip = 'NA'
                    if 'ENDBREAKSEQ' in formatdict:
                        leftsnip = formatdict['ENDBREAKSEQ']

                    leftconskey  = ':'.join((chrom, pos, alt, '1'))
                    rightconskey = ':'.join((chrom, pos, alt, '2'))

                    leftcons = 'NA'
                    if leftconskey in consdict:
                        leftcons = consdict[leftconskey]

                    rightcons = 'NA'
                    if rightconskey in consdict:
                        rightcons = consdict[rightconskey]

                    liblist  = formatdict['RG']
                    libcount = len(liblist.split(','))
                    mapscore = 'NA'
                    if mapscore in infodict:
                        mapscore = infodict['MAPSCORE']

                    # annotation stubs, will be filled in at summary
                    refelt = 'NA' 
                    nonref = 'NA' 
                    genes  = 'NA'
                    reg    = 'NA'

                    data = [chrom, start, end, posconf, strand, strandconf, alt, family, famconf, refelt,
                            eltlen, teminpos, temaxpos, TSD, tlen, DEL, dlen, rightsnip, leftsnip, leftcons,
                            rightcons, lsupport, rsupport, support, libcount, liblist, mapscore, nonref, 
                            genes, reg]

                    assert len(data) == len(columns), "column count: " + str(len(columns)) + " data count: " + str(len(data))

                    data = map(str, data)
                    for i in range(len(data)):
                        tld[chrom + ':' + pos][columns[i]] = data[i]

    return tld


def parseGFF(chrom, start, end, gffhandle):
    featurelist = ['gene', 'exon', 'CDS', 'UTR']
    annotations = [] 
    if chrom in gffhandle.contigs:
        for rec in gffhandle.fetch(chrom, int(start)-1, int(end)+1):
            chrom, source, feature, start, end, score, strand, frame = rec.strip().split()[:8]
            attribs = [attrib.strip().replace('"','') for attrib in ' '.join(rec.strip().split()[8:]).split(';')]
            gname = 'NA'
            for attrib in attribs:
                if attrib.startswith('gene_name'):
                    gname = '_'.join(attrib.split()[1:])
            if feature in featurelist:
                annotations.append('|'.join((feature, gname, chrom, start, end, strand)))
    return annotations


def parseBED(chrom, start, end, bedhandle):
    annotations = []
    if chrom in bedhandle.contigs:
        for rec in bedhandle.fetch(chrom, int(start)-1, int(end)+1):
            chrom, start, end, info = rec.strip().split()[:4]
            annotations.append(info)

    return list(set(annotations))


def getlibcounts(libstring):
    ''' libstring is name1|count1,name2|count2,... return dictionary of library name --> count '''
    return dict([(l,int(c)) for l,c in [lc.split('|') for lc in libstring.split(',')]])


def checkexclude(excludelibs, libcounts, maxexfrac=0.1):
    ''' libcounts is a dictionary libname --> supporting read count '''
    ''' return True if filtered due to > maxexfrac reads supporting library in excludelibs '''
    exctotal = 0.0
    alltotal = 0.0
    if excludelibs is not None:
        for lib, count in libcounts.iteritems():
            if lib in excludelibs:
                exctotal += float(count)
            alltotal += float(count)

        if exctotal/alltotal > maxexfrac:
            return True
        return False

    else:
        return False
            

def summary(tldlist, wb, args, mastertld=None, tophits=False, excludelibs=None):
    ''' sequential merge of TLDs (two-level dictionaries) '''
    if mastertld is None:
        mastertld = tldlist[0]
        if len(tldlist) > 1:
            for tld in tldlist[1:]:
                mastertld = mergetlds(mastertld, tld)

    # init Tabix handles
    if args.ref is not None:
        reftbx = Tabixfile(args.ref)
    if args.nrtabix is not None:
        nr = Tabixfile(args.nrtabix)
    if args.bed is not None:
        bedhandle = Tabixfile(args.bed)
    if args.gff is not None:
        gffhandle = Tabixfile(args.gff)
    if args.mapscore is not None:
        maptabix=Tabixfile(args.mapscore)

    rmsktabix = Tabixfile(args.ref)

    # Annotation
    if not tophits:
        print "Annotating summary ... "
        for insloc in mastertld.keys():
            assert mastertld[insloc]['Ref_TE'] == 'NA'
            assert mastertld[insloc]['Known_Nonref'] == 'NA'
            assert mastertld[insloc]['Genes'] == 'NA'
            assert mastertld[insloc]['Regulation'] == 'NA'

            chrom = mastertld[insloc]['Chr']
            start = int(mastertld[insloc]['Left_Position'])
            end   = int(mastertld[insloc]['Right_Position'])

            # 'Ref_TE'
            refwindow = int(args.refwindow)
            if args.ref is not None:
                if chrom in reftbx.contigs:
                    for rec in reftbx.fetch(chrom, start-refwindow, end+refwindow):
                        refclass = rec.split()[3]
                        if refclass == mastertld[insloc]['Class']:
                            mastertld[insloc]['Ref_TE'] = '|'.join(rec.strip().split())

            # 'Known_Nonref'
            searchwindow = 100
            if args.nrtabix is not None:
                if chrom in nr.contigs:
                    nonreflist = [elt.split()[3] + '|' + elt.split()[-1] for elt in nr.fetch(chrom, start-searchwindow, end+searchwindow)]
                    if nonreflist:
                        mastertld[insloc]['Known_Nonref'] = ','.join(nonreflist)

            # 'Genes'
            if args.gff is not None:
                mastertld[insloc]['Genes'] = ','.join(parseGFF(chrom, start, end, gffhandle))
            if not mastertld[insloc]['Genes']:
                mastertld[insloc]['Genes'] = 'NA'

            # 'Regulation'
            if args.bed is not None:
                mastertld[insloc]['Regulation'] = ','.join(parseBED(chrom, start, end, bedhandle))
            if not mastertld[insloc]['Regulation']:
                mastertld[insloc]['Regulation'] = 'NA'

    # Filtering
    filtered = {}
    for insloc in mastertld.keys():
        assert len(mastertld[insloc]) > 0, "Error: zero length insloc: " + insloc 
        if mastertld[insloc]['Ref_TE'] != 'NA':
            filtered[insloc] = True

        if int(mastertld[insloc]['Total_Support']) < int(args.minsupport):
            filtered[insloc] = True

        if tophits:
            if not args.allowknown:
                if mastertld[insloc]['Known_Nonref'] != 'NA':
                    filtered[insloc] = True

            libcounts = getlibcounts(mastertld[insloc]['Library_List'])

            if checkexclude(excludelibs, libcounts):
                filtered[insloc] = True

        ldata = {}
        rdata = {}
        if insloc not in filtered and tophits: # only run BLAT filter for 'tophits'
            if 'Bad_BLAT' not in columns:
                columns.append('Bad_BLAT')

            if 'BLAT_Filter_Data_Left' not in columns:
                columns.append('BLAT_Filter_Data_Left')

            if 'BLAT_Filter_Data_Right' not in columns:
                columns.append('BLAT_Filter_Data_Right')

            if mastertld[insloc]['Left_Consensus'] != 'NA':
                chrom = mastertld[insloc]['Chr']
                lpos  = mastertld[insloc]['Left_Position']
                lcons = mastertld[insloc]['Left_Consensus']
                ldata = blatfilter.checkseq(lcons, chrom, lpos, mastertld[insloc]['Class'], rmsktabix, args.genomeref, args.teref, args.refport, args.teport, maptabix=maptabix, tlfilter=args.tlfilter)
                mastertld[insloc]['BLAT_Filter_Data_Left'] = str(ldata)
                mastertld[insloc]['BLAT_Filter_Data_Right'] = 'NA' # default


            if mastertld[insloc]['Right_Consensus'] != 'NA':
                chrom = mastertld[insloc]['Chr']
                rpos  = mastertld[insloc]['Right_Position']
                rcons = mastertld[insloc]['Right_Consensus']
                rdata = blatfilter.checkseq(rcons, chrom, rpos, mastertld[insloc]['Class'], rmsktabix, args.genomeref, args.teref, args.refport, args.teport, maptabix=maptabix, tlfilter=args.tlfilter)
                mastertld[insloc]['BLAT_Filter_Data_Right'] = str(rdata)

            badblat = True
            if 'pass' in ldata and ldata['pass']:
                badblat = False

            if 'pass' in rdata and rdata['pass']:
                badblat = False

            mastertld[insloc]['Bad_BLAT'] = str(badblat)

    for insloc in filtered.keys():
        del mastertld[insloc]

    if tophits:
        for insloc in mastertld.keys():
            if 'Bad_BLAT' in mastertld[insloc] and mastertld[insloc]['Bad_BLAT'] == 'True':
                del mastertld[insloc]

    ws = None
 
    if tophits:
        ws = wb.create_sheet(0)
        ws.title="TopHits"
    else: 
        ws = wb.create_sheet(0)
        ws.title="Summary"

    assert ws is not None

    current_row = 0
    for i in range(len(columns)):
        ws.cell(row=current_row, column=i).value = columns[i]
    current_row += 1

    for insrow in mastertld.keys():
        data = []
        for c in columns:
            if c not in mastertld[insrow]:
                mastertld[insrow][c] = 'BUG'
            data.append(mastertld[insrow][c])

        data = map(str, data)
        for i in range(len(data)):
            ws.cell(row=current_row, column=i).value = data[i]
        current_row += 1
        
    return wb, mastertld


def mergetlds(tld1, tld2):
    newtld = dd(dict)

    for insloc in tld1.keys():
        if insloc in tld2:
            # check for class mismatch
            if tld1[insloc]['Class'] != tld2[insloc]['Class']:
                sys.stderr.write("WARNING: Element class mismatch at " + insloc + "\n")
                newtld[insloc] = tld1[insloc]

            else:
                # merge
                newtld[insloc]['Chr'] = tld1[insloc]['Chr']

                default   = tld1
                best_left = tld1
                best_right = tld1

                if int(tld2[insloc]['Left_Support']) > int(tld1[insloc]['Left_Support']):
                    best_left = tld2

                if int(tld2[insloc]['Right_Support']) > int(tld1[insloc]['Right_Support']):
                    best_right = tld2

                best_strand = tld1
                if float(tld2[insloc]['Strand_Conf']) > float(tld1[insloc]['Strand_Conf']):
                    best_strand = tld2

                best_family = tld1
                if float(tld2[insloc]['Family_Conf']) > float(tld1[insloc]['Family_Conf']):
                    best_family = tld2

                newlength = tld1[insloc]['Elt_Length']
                if newlength == 'NA' and tld2[insloc]['Elt_Length'] != 'NA':
                    newlength = tld2[insloc]['Elt_Length'] != 'NA'
                
                if 'NA' not in (tld1[insloc]['Elt_Length'], tld2[insloc]['Elt_Length']):
                    newlength = max(int(tld1[insloc]['Elt_Length']), int(tld2[insloc]['Elt_Length']))

                best_tsd = tld1
                best_del = tld1
                if tld1[insloc]['TSD'] == 'N' and tld2[insloc]['TSD'] == 'Y':
                    best_tsd = tld2

                if tld1[insloc]['DEL'] == 'N' and tld2[insloc]['DEL'] == 'Y':
                    best_del = tld2

                total_lsupport = int(tld1[insloc]['Left_Support']) + int(tld2[insloc]['Left_Support'])
                total_rsupport = int(tld1[insloc]['Right_Support']) + int(tld2[insloc]['Right_Support'])

                combined_libcount = int(tld1[insloc]['Libraries']) + int(tld2[insloc]['Libraries'])
                combined_liblist  = tld1[insloc]['Library_List'] + ',' + tld2[insloc]['Library_List'] 

                newtld[insloc]['Chr']             = best_left[insloc]['Chr']
                newtld[insloc]['Left_Position']   = best_left[insloc]['Left_Position']
                newtld[insloc]['Right_Position']  = best_right[insloc]['Right_Position']
                newtld[insloc]['Strand']          = best_strand[insloc]['Strand']
                newtld[insloc]['Strand_Conf']     = best_strand[insloc]['Strand_Conf']
                newtld[insloc]['Class']           = best_family[insloc]['Class']
                newtld[insloc]['Family']          = best_family[insloc]['Family']
                newtld[insloc]['Family_Conf']     = best_family[insloc]['Family_Conf']
                newtld[insloc]['Ref_TE']          = default[insloc]['Ref_TE']
                newtld[insloc]['Elt_Length']      = newlength
                newtld[insloc]['TSD']             = best_tsd[insloc]['TSD']
                newtld[insloc]['TSD_Length']      = best_tsd[insloc]['TSD_Length']
                newtld[insloc]['DEL']             = best_del[insloc]['DEL']
                newtld[insloc]['DEL_Length']      = best_del[insloc]['DEL_Length']
                newtld[insloc]['Left_Snip']       = best_left[insloc]['Left_Snip']
                newtld[insloc]['Right_Snip']      = best_left[insloc]['Right_Snip']
                newtld[insloc]['Left_Consensus']  = best_left[insloc]['Left_Consensus'] 
                newtld[insloc]['Right_Consensus'] = best_left[insloc]['Right_Consensus'] 
                newtld[insloc]['Left_Support']    = total_lsupport
                newtld[insloc]['Right_Support']   = total_rsupport
                newtld[insloc]['Total_Support']   = total_lsupport + total_rsupport
                newtld[insloc]['Libraries']       = combined_libcount
                newtld[insloc]['Library_List']    = combined_liblist
                newtld[insloc]['Mapscore']        = default[insloc]['Mapscore']
                newtld[insloc]['Known_Nonref']    = default[insloc]['Known_Nonref']
                newtld[insloc]['Genes']           = default[insloc]['Genes']
                newtld[insloc]['Regulation']      = default[insloc]['Regulation']

        else:
            newtld[insloc] = tld1[insloc]

        # add all tld2 not in tld1
        for insloc in tld2.keys():
            if insloc not in newtld:
                newtld[insloc] = tld2[insloc]

    return newtld


def main(args):
    wb = Workbook()
    tldlist = []

    assert args.genomeref.endswith('2bit')
    assert args.teref.endswith('2bit')

    exlibs = None
    if args.excludelibs is not None:
        with open(args.excludelibs, 'r') as exfile:
            exlibs = [line.strip() for line in exfile]

    with open(args.indirlist, 'r') as indirlist:
        for indir in indirlist:
            indir = indir.strip()
            sname = os.path.basename(indir)
            invcf = indir + '/' + sname + '.vcf'
            consfasta = indir + '/' + sname + '.cons.fa'

            assert os.path.exists(invcf), "missing file: " + invcf
            assert os.path.exists(consfasta), "missing file: " + consfasta

            tld = vcf2tld(args, invcf, consfasta)
            tldlist.append(tld)

    p = blatfilter.start_blat_server(args.genomeref, port=args.refport)
    t = blatfilter.start_blat_server(args.teref, port=args.teport)

    try:
        wb, sumtld = summary(tldlist, wb, args)
        wb, toptld = summary(tldlist, wb, args, mastertld=sumtld, tophits=True, excludelibs=exlibs)
        wb.save(args.out)

    except Exception, e:
        sys.stderr.write("*"*60 + "\nerror in blat filter:\n")
        traceback.print_exc(file=sys.stderr)
        sys.stderr.write("*"*60 + "\n")

    print "killing BLAT server(s) ..."
    p.kill()
    t.kill()


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='make a human-readable table from tebreak VCF')
    parser.add_argument('-i', '--indirlist', required=True, help='list of tebreak output directories used as input')
    parser.add_argument('-r', '--ref', default=None, help='reference element locations tabix, column 4 should correspond to element class')
    parser.add_argument('-n', '--nrtabix', default=None, help='known non-ref element tabix')
    parser.add_argument('-o', '--out', default='out.xlsx', help='output excel')
    parser.add_argument('-m', '--mapscore', default=None, help='mapscore tabix')
    parser.add_argument('--gff', default=None, help='GFF annotation file (genes, exons)')
    parser.add_argument('--bed', default=None, help='BED annotation file (regulatory info)')
    parser.add_argument('--refwindow', default=50, help='distance cutoff for finding reference elements')
    parser.add_argument('--excludelibs', default=None, help='file containing list of libraries to exclude')
    parser.add_argument('--genomeref', required=True, help='genome BLAT reference (2bit)')
    parser.add_argument('--teref', required=True, help='TE BLAT reference (2bit)')
    parser.add_argument('--minsupport', default=5, help='minimum support (read count, default=5)')
    parser.add_argument('--tlfilter', action='store_true', help='turn on translocation filter (stringent/experimental)')
    parser.add_argument('--allowknown', action='store_true', help='keep known non-reference insertions even if other filters exclude them')
    parser.add_argument('--refport', default=9999)
    parser.add_argument('--teport', default=9998)
    args = parser.parse_args()
    main(args)
